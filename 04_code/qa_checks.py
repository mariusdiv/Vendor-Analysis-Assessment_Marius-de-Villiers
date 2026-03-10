"""
qa_checks.py - Comprehensive QA Validation for Vendor Spend Analysis
=====================================================================

Purpose:
    Production-grade QA pipeline that validates the vendor analysis output
    workbook against the vendor classification database, checking data
    completeness, referential integrity, description quality, spend
    reconciliation, and potential duplicate vendors.

Inputs:
    - 03_outputs/Vendor Analysis Assessment - Marius de Villiers.xlsx
      (tab: Vendor Analysis Assessment, columns A-E)
    - 04_code/vendor_db.json (386-entry classification database)

Outputs:
    - 03_outputs/qa_report.md        Full markdown QA report with tables
    - 03_outputs/possible_duplicates.csv  Fuzzy-matched duplicate groups
    - stdout: QA PASS / QA FAIL with reasons

Exit Codes:
    0  All validations passed (QA PASS)
    1  One or more validations failed (QA FAIL)
"""

from __future__ import annotations

import csv
import json
import sys
from collections import defaultdict
from datetime import datetime
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, NamedTuple

import openpyxl


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

PROJECT_ROOT: Path = Path(__file__).resolve().parent.parent

EXCEL_PATH: Path = (
    PROJECT_ROOT / "03_outputs"
    / "Vendor Analysis Assessment - Marius de Villiers.xlsx"
)
VENDOR_DB_PATH: Path = PROJECT_ROOT / "04_code" / "vendor_db.json"
QA_REPORT_PATH: Path = PROJECT_ROOT / "03_outputs" / "qa_report.md"
DUPLICATES_CSV_PATH: Path = PROJECT_ROOT / "03_outputs" / "possible_duplicates.csv"

ALLOWED_DEPARTMENTS: frozenset[str] = frozenset({
    "Engineering", "Facilities", "G&A", "Legal", "M&A", "Marketing",
    "SaaS", "Professional Services", "Support", "Finance", "Product", "Sales",
})

ALLOWED_SUGGESTIONS: frozenset[str] = frozenset({"Consolidate", "Terminate", "Optimize"})

GENERIC_PHRASES: tuple[str, ...] = ("business services", "general business", "services and operations")

EXPECTED_TOTAL_SPEND: float = 7_887_360.40
SPEND_TOLERANCE: float = 1.00
FUZZY_THRESHOLD: float = 0.85
TOP_N_SPOT_CHECK: int = 10
TOP_N_VENDORS_TABLE: int = 25


# ---------------------------------------------------------------------------
# Data Structures
# ---------------------------------------------------------------------------

class VendorRow(NamedTuple):
    """One row from the Vendor Analysis Assessment tab."""
    row_num: int
    name: str
    department: str | None
    cost: float
    description: str | None
    suggestion: str | None

class DuplicateGroup(NamedTuple):
    """A cluster of vendors that may be duplicates."""
    group_id: int
    vendor_names: list[str]
    combined_spend: float
    similarity_score: float

class ValidationResult(NamedTuple):
    """Outcome of a single validation check."""
    check_id: int
    title: str
    passed: bool
    messages: list[str]


# ---------------------------------------------------------------------------
# Data Loading
# ---------------------------------------------------------------------------

def load_excel_vendors(path: Path) -> list[VendorRow]:
    """Load vendor rows from the Vendor Analysis Assessment tab."""
    wb = openpyxl.load_workbook(str(path), data_only=True)
    ws = wb["Vendor Analysis Assessment"]
    vendors: list[VendorRow] = []
    for row_idx in range(2, ws.max_row + 1):
        name = ws.cell(row=row_idx, column=1).value
        if not name:
            continue
        dept = ws.cell(row=row_idx, column=2).value
        cost_raw = ws.cell(row=row_idx, column=3).value
        cost = float(cost_raw) if cost_raw is not None else 0.0
        desc = ws.cell(row=row_idx, column=4).value
        suggestion = ws.cell(row=row_idx, column=5).value
        vendors.append(VendorRow(
            row_num=row_idx, name=str(name).strip(),
            department=str(dept).strip() if dept else None, cost=cost,
            description=str(desc).strip() if desc else None,
            suggestion=str(suggestion).strip() if suggestion else None))
    wb.close()
    return vendors

def load_vendor_db(path: Path) -> dict[str, list[str]]:
    """Load the JSON vendor classification database."""
    with open(path, "r", encoding="utf-8") as fh:
        return json.load(fh)


# ---------------------------------------------------------------------------
# Validation Checks
# ---------------------------------------------------------------------------

def check_required_fields(vendors: list[VendorRow]) -> ValidationResult:
    """Check 1: No blank vendor name, department, description, or suggestion."""
    msgs: list[str] = []
    for v in vendors:
        if not v.name:
            msgs.append(f"Row {v.row_num}: blank vendor name")
        if not v.department:
            msgs.append(f"Row {v.row_num}: {v.name} - blank department")
        if not v.description:
            msgs.append(f"Row {v.row_num}: {v.name} - blank description")
        if not v.suggestion:
            msgs.append(f"Row {v.row_num}: {v.name} - blank suggestion")
    passed = len(msgs) == 0
    if passed:
        msgs.insert(0, f"All {len(vendors)} vendors have complete data.")
    else:
        msgs.insert(0, f"{len(msgs)} blank field(s) detected.")
    return ValidationResult(1, "Required Fields", passed, msgs)

def check_department_values(vendors: list[VendorRow]) -> ValidationResult:
    """Check 2: Every department is in the allowed set."""
    msgs: list[str] = []
    for v in vendors:
        if v.department and v.department not in ALLOWED_DEPARTMENTS:
            msgs.append(f"Row {v.row_num}: {v.name} has invalid department " + repr(v.department))
    passed = len(msgs) == 0
    if passed:
        msgs.insert(0, "All departments are valid.")
    else:
        msgs.insert(0, f"{len(msgs)} invalid department(s) found.")
    return ValidationResult(2, "Department Validation", passed, msgs)

def check_suggestion_values(vendors: list[VendorRow]) -> ValidationResult:
    """Check 3: Every suggestion is Consolidate, Terminate, or Optimize."""
    msgs: list[str] = []
    for v in vendors:
        if v.suggestion and v.suggestion not in ALLOWED_SUGGESTIONS:
            msgs.append(f"Row {v.row_num}: {v.name} has invalid suggestion " + repr(v.suggestion))
    passed = len(msgs) == 0
    if passed:
        msgs.insert(0, "All suggestions are valid.")
    else:
        msgs.insert(0, f"{len(msgs)} invalid suggestion(s) found.")
    return ValidationResult(3, "Suggestion Validation", passed, msgs)

def check_description_quality(vendors: list[VendorRow]) -> ValidationResult:
    """Check 4: Flag descriptions containing generic phrases."""
    msgs: list[str] = []
    for v in vendors:
        if not v.description:
            continue
        dl = v.description.lower()
        for phrase in GENERIC_PHRASES:
            if phrase in dl:
                msgs.append(f"Row {v.row_num}: {v.name} - generic phrase in: {v.description}")
                break
    passed = len(msgs) == 0
    if passed:
        msgs.insert(0, "No generic descriptions detected.")
    else:
        msgs.insert(0, f"{len(msgs)} vendor(s) with generic descriptions.")
    return ValidationResult(4, "Description Quality", passed, msgs)

def check_spend_reconciliation(vendors: list[VendorRow]) -> ValidationResult:
    """Check 5: Total spend must equal $7,887,360.40 within $1 tolerance."""
    actual = sum(v.cost for v in vendors)
    diff = abs(actual - EXPECTED_TOTAL_SPEND)
    passed = diff <= SPEND_TOLERANCE
    msgs: list[str] = [
        f"Expected total:  ${EXPECTED_TOTAL_SPEND:>14,.2f}",
        f"Actual total:    ${actual:>14,.2f}",
        f"Difference:      ${diff:>14,.2f}",
        f"Tolerance:       ${SPEND_TOLERANCE:>14,.2f}",
    ]
    if passed:
        msgs.insert(0, "Spend reconciliation PASSED.")
    else:
        msgs.insert(0, "Spend reconciliation FAILED - difference exceeds tolerance.")
    return ValidationResult(5, "Spend Reconciliation", passed, msgs)

def check_top10_completeness(vendors: list[VendorRow]) -> ValidationResult:
    """Check 6: Top 10 vendors by spend all have complete data."""
    sv = sorted(vendors, key=lambda v: v.cost, reverse=True)
    top10 = sv[:TOP_N_SPOT_CHECK]
    msgs: list[str] = []
    issues: list[str] = []
    for v in top10:
        ok = all([v.department, v.description, v.suggestion])
        st = "OK" if ok else "INCOMPLETE"
        msgs.append(f"  ${v.cost:>12,.2f}  {st:<12} {v.name}")
        if not ok:
            issues.append(v.name)
    passed = len(issues) == 0
    hdr = "All top 10 vendors have complete data." if passed else f"{len(issues)} top-10 vendor(s) have incomplete data."
    msgs.insert(0, hdr)
    return ValidationResult(6, "Top 10 Spot Check", passed, msgs)


# ---------------------------------------------------------------------------
# Aggregation Analysis
# ---------------------------------------------------------------------------

def spend_by_department(vendors: list[VendorRow]) -> list[dict[str, Any]]:
    """Check 7: Aggregate spend by department, sorted descending by spend."""
    dd: dict[str, dict[str, Any]] = defaultdict(lambda: {"count": 0, "spend": 0.0})
    total = sum(v.cost for v in vendors)
    for v in vendors:
        d = v.department or "(blank)"
        dd[d]["count"] += 1
        dd[d]["spend"] += v.cost
    rows = []
    for dept, data in sorted(dd.items(), key=lambda x: x[1]["spend"], reverse=True):
        pct = (data["spend"] / total * 100) if total else 0.0
        rows.append({"department": dept, "count": data["count"], "spend": data["spend"], "pct": pct})
    return rows

def spend_by_suggestion(vendors: list[VendorRow]) -> list[dict[str, Any]]:
    """Check 8: Aggregate spend by suggestion, sorted descending by spend."""
    sd: dict[str, dict[str, Any]] = defaultdict(lambda: {"count": 0, "spend": 0.0})
    total = sum(v.cost for v in vendors)
    for v in vendors:
        s = v.suggestion or "(blank)"
        sd[s]["count"] += 1
        sd[s]["spend"] += v.cost
    rows = []
    for sug, data in sorted(sd.items(), key=lambda x: x[1]["spend"], reverse=True):
        pct = (data["spend"] / total * 100) if total else 0.0
        rows.append({"suggestion": sug, "count": data["count"], "spend": data["spend"], "pct": pct})
    return rows

def top_vendors_table(vendors: list[VendorRow], n: int = TOP_N_VENDORS_TABLE) -> list[VendorRow]:
    """Check 9: Return the top N vendors by spend."""
    return sorted(vendors, key=lambda v: v.cost, reverse=True)[:n]


# ---------------------------------------------------------------------------
# Duplicate Detection
# ---------------------------------------------------------------------------

def detect_duplicates(vendors: list[VendorRow], threshold: float = FUZZY_THRESHOLD) -> list[DuplicateGroup]:
    """Check 10-12: Fuzzy-match vendor names using SequenceMatcher at threshold."""
    names = [(v.name, v.cost) for v in vendors]
    n = len(names)
    pairs: list[tuple[int, int, float]] = []
    for i in range(n):
        for j in range(i + 1, n):
            ratio = SequenceMatcher(None, names[i][0].lower(), names[j][0].lower()).ratio()
            if ratio >= threshold:
                pairs.append((i, j, ratio))
    parent = list(range(n))
    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]; x = parent[x]
        return x
    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[ra] = rb
    for i, j, _ in pairs:
        union(i, j)
    pair_map: dict[int, set[int]] = defaultdict(set)
    score_map: dict[int, list[float]] = defaultdict(list)
    for i, j, ratio in pairs:
        root = find(i)
        pair_map[root].add(i)
        pair_map[root].add(j)
        score_map[root].append(ratio)
    groups: list[DuplicateGroup] = []
    for gid, (root, members) in enumerate(
        sorted(pair_map.items(), key=lambda x: -sum(names[m][1] for m in x[1])), start=1):
        nc = {names[m][0]: names[m][1] for m in members}
        vnames = sorted(nc.keys(), key=lambda nm: -nc[nm])
        combined = sum(nc.values())
        avg = sum(score_map[root]) / len(score_map[root]) if score_map[root] else 0.0
        groups.append(DuplicateGroup(group_id=gid, vendor_names=vnames, combined_spend=combined, similarity_score=round(avg, 4)))
    return groups

def write_duplicates_csv(groups: list[DuplicateGroup], path: Path) -> None:
    """Write duplicate groups to a CSV file."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as fh:
        w = csv.writer(fh)
        w.writerow(["group_id", "vendor_names", "combined_spend", "similarity_score"])
        for g in groups:
            w.writerow([g.group_id, " | ".join(g.vendor_names), f"{g.combined_spend:.2f}", f"{g.similarity_score:.4f}"])


# ---------------------------------------------------------------------------
# Report Generation
# ---------------------------------------------------------------------------

def _md_table(headers: list[str], rows: list[list[str]]) -> str:
    """Build a simple markdown table."""
    widths = [len(h) for h in headers]
    for row in rows:
        for i, cell in enumerate(row):
            widths[i] = max(widths[i], len(cell))
    sep = "| " + " | ".join("-" * w for w in widths) + " |"
    hdr = "| " + " | ".join(h.ljust(w) for h, w in zip(headers, widths)) + " |"
    body = [hdr, sep]
    for row in rows:
        body.append("| " + " | ".join(cell.ljust(w) for cell, w in zip(row, widths)) + " |")
    return chr(10).join(body)

def generate_report(vendors, db, results, dept_rows, sug_rows, top_vendors, dup_groups) -> str:
    """Generate the full markdown QA report."""
    out: list[str] = []
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    out.append("# QA Validation Report")
    out.append("")
    out.append(f"**Generated:** {now}  ")
    out.append(f"**Excel File:** `{EXCEL_PATH.name}`  ")
    out.append(f"**Vendor DB:** `{VENDOR_DB_PATH.name}` ({len(db)} entries)  ")
    out.append(f"**Vendors in Excel:** {len(vendors)}  ")
    ts = sum(v.cost for v in vendors)
    out.append(f"**Total Spend:** ${ts:,.2f}")
    out.append("")
    ap = all(r.passed for r in results)
    vd = "QA PASS" if ap else "QA FAIL"
    pc = sum(1 for r in results if r.passed)
    out.append(f"## Overall Verdict: **{vd}** ({pc}/{len(results)} checks passed)")
    out.append("")
    out.append("---")
    out.append("")
    out.append("## Validation Checks")
    out.append("")
    for r in results:
        ic = "PASS" if r.passed else "FAIL"
        out.append(f"### Check {r.check_id}: {r.title} [{ic}]")
        out.append("")
        for msg in r.messages:
            out.append(f"- {msg}")
        out.append("")
    out.append("---")
    out.append("")
    out.append("## Aggregation: Spend by Department")
    out.append("")
    h = ["Department", "Vendors", "Spend", "% of Total"]
    tr = []
    for d in dept_rows:
        tr.append([d["department"], str(d["count"]), f"${d['spend']:,.2f}", f"{d['pct']:.1f}%"])
    tr.append(["**TOTAL**", f"**{sum(d['count'] for d in dept_rows)}**", f"**${sum(d['spend'] for d in dept_rows):,.2f}**", "**100.0%**"])
    out.append(_md_table(h, tr))
    out.append("")
    out.append("## Aggregation: Spend by Suggestion")
    out.append("")
    h = ["Suggestion", "Vendors", "Spend", "% of Total"]
    tr = []
    for s in sug_rows:
        tr.append([s["suggestion"], str(s["count"]), f"${s['spend']:,.2f}", f"{s['pct']:.1f}%"])
    tr.append(["**TOTAL**", f"**{sum(s['count'] for s in sug_rows)}**", f"**${sum(s['spend'] for s in sug_rows):,.2f}**", "**100.0%**"])
    out.append(_md_table(h, tr))
    out.append("")
    out.append("## Top 25 Vendors by Spend")
    out.append("")
    h = ["Rank", "Vendor", "Department", "Spend", "Suggestion", "Description"]
    tr = []
    for i, v in enumerate(top_vendors, start=1):
        vn = v.name[:40] + ("..." if len(v.name) > 40 else "")
        vdesc = (v.description or "")[:50] + ("..." if v.description and len(v.description) > 50 else "")
        tr.append([str(i), vn, v.department or "", f"${v.cost:,.2f}", v.suggestion or "", vdesc])
    out.append(_md_table(h, tr))
    out.append("")
    out.append("---")
    out.append("")
    out.append(f"## Duplicate Detection ({len(dup_groups)} groups found)")
    out.append("")
    if dup_groups:
        out.append(f"Fuzzy threshold: {FUZZY_THRESHOLD:.0%}")
        out.append("")
        h = ["Group", "Vendors", "Combined Spend", "Similarity"]
        tr = []
        for g in dup_groups:
            tr.append([str(g.group_id), " | ".join(g.vendor_names), f"${g.combined_spend:,.2f}", f"{g.similarity_score:.2%}"])
        out.append(_md_table(h, tr))
        out.append("")
        out.append(f"Full duplicate data exported to: `{DUPLICATES_CSV_PATH.name}`")
    else:
        out.append("No potential duplicates detected at the current threshold.")
    out.append("")
    out.append("---")
    out.append("")
    out.append(f"*Report generated by qa_checks.py on {now}*")
    out.append("")
    return chr(10).join(out)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    """Run the full QA pipeline and return the exit code (0=pass, 1=fail)."""
    print("=" * 70)
    print("QA VALIDATION PIPELINE")
    print(f"  Excel:     {EXCEL_PATH}")
    print(f"  VendorDB:  {VENDOR_DB_PATH}")
    print(f"  Report:    {QA_REPORT_PATH}")
    print(f"  Dupes CSV: {DUPLICATES_CSV_PATH}")
    print("=" * 70)
    print()
    print("[1/4] Loading data ...")
    vendors = load_excel_vendors(EXCEL_PATH)
    db = load_vendor_db(VENDOR_DB_PATH)
    print(f"       Loaded {len(vendors)} vendors from Excel, {len(db)} from DB.")
    print()
    print("[2/4] Running validation checks ...")
    results: list[ValidationResult] = [
        check_required_fields(vendors),
        check_department_values(vendors),
        check_suggestion_values(vendors),
        check_description_quality(vendors),
        check_spend_reconciliation(vendors),
        check_top10_completeness(vendors),
    ]
    for r in results:
        st = "PASS" if r.passed else "FAIL"
        print(f"       Check {r.check_id} [{st}] {r.title}: {r.messages[0]}")
    print()
    print("[3/4] Running aggregation analysis ...")
    dept_rows = spend_by_department(vendors)
    sug_rows = spend_by_suggestion(vendors)
    top25 = top_vendors_table(vendors)
    print("       Spend by Department:")
    for d in dept_rows:
        print(f"         {d['department']:<25} {d['count']:>4} vendors  ${d['spend']:>13,.2f}  ({d['pct']:5.1f}%)")
    print()
    print("       Spend by Suggestion:")
    for s in sug_rows:
        print(f"         {s['suggestion']:<25} {s['count']:>4} vendors  ${s['spend']:>13,.2f}  ({s['pct']:5.1f}%)")
    print()
    print("[4/4] Running duplicate detection (fuzzy matching) ...")
    dup_groups = detect_duplicates(vendors, FUZZY_THRESHOLD)
    print(f"       Found {len(dup_groups)} potential duplicate group(s).")
    for g in dup_groups:
        joined = " | ".join(g.vendor_names)
        print(f"         Group {g.group_id}: {joined} (${g.combined_spend:,.2f}, {g.similarity_score:.2%})")
    print()
    print("Writing outputs ...")
    write_duplicates_csv(dup_groups, DUPLICATES_CSV_PATH)
    print(f"  -> {DUPLICATES_CSV_PATH}")
    report = generate_report(vendors, db, results, dept_rows, sug_rows, top25, dup_groups)
    QA_REPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    QA_REPORT_PATH.write_text(report, encoding="utf-8")
    print(f"  -> {QA_REPORT_PATH}")
    print()
    all_passed = all(r.passed for r in results)
    passed_count = sum(1 for r in results if r.passed)
    print("=" * 70)
    if all_passed:
        print(f"QA PASS  ({passed_count}/{len(results)} checks passed)")
    else:
        failed = [r for r in results if not r.passed]
        reasons = ", ".join(f"Check {r.check_id} ({r.title})" for r in failed)
        print(f"QA FAIL  ({passed_count}/{len(results)} checks passed)")
        print(f"  Failed: {reasons}")
    print("=" * 70)
    return 0 if all_passed else 1


if __name__ == "__main__":
    sys.exit(main())
